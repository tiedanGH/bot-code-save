# save as img_to_excel.py
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def image_to_excel(img_path, out_xlsx="pixel_art.xlsx", background="#FFFFFF", max_side=None):
    """
    将图片每个像素写入 Excel 单元格背景色。
    - img_path: 图片路径
    - out_xlsx: 输出文件名
    - background: 若图片有透明，先与该背景合成。格式如 "#FFFFFF" 或 "black"
    - max_side: 若不为 None，则最长边缩放到该值，保持纵横比
    """
    # 打开并统一为 RGBA（保留透明信息以便合成）
    img = Image.open(img_path).convert("RGBA")
    w, h = img.size

    # 可选缩放（保持比例）
    if max_side is not None:
        scale = min(1.0, max_side / max(w, h))
        if scale < 1.0:
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            img = img.resize(new_size, Image.LANCZOS)
            w, h = img.size

    print(f"图片大小: {w} x {h}")

    # 处理背景颜色字符串 -> RGB tuple
    if isinstance(background, str) and background.startswith("#"):
        bg = tuple(int(background[i:i+2], 16) for i in (1, 3, 5))
    elif isinstance(background, str) and background.lower() in ("white", "#fff", "#ffffff"):
        bg = (255, 255, 255)
    elif isinstance(background, str) and background.lower() in ("black", "#000", "#000000"):
        bg = (0, 0, 0)
    else:
        # 默认白色
        bg = (255, 255, 255)

    # 如果有透明通道，先把图片合成到指定背景上
    if img.mode == "RGBA":
        background_img = Image.new("RGBA", img.size, bg + (255,))
        img = Image.alpha_composite(background_img, img).convert("RGB")
    else:
        img = img.convert("RGB")

    wb = Workbook()
    ws = wb.active

    # 为兼容性，使用 AARRGGBB 格式（FF 表示不透明）
    for y in range(h):
        for x in range(w):
            r, g, b = img.getpixel((x, y))
            hex_color = f"FF{r:02X}{g:02X}{b:02X}"  # AARRGGBB
            fill = PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)
            ws.cell(row=y+1, column=x+1).fill = fill

    # 调整列宽和行高以尽量接近正方形像素
    # 下面的值可以根据你的 Excel 版本微调
    col_width = 2.1
    row_height = 9  # 注意：不同系统、不同 Excel 行高单位不同，调到看起来正方形为止
    for col_idx in range(1, w+1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_width
    for row_idx in range(1, h+1):
        ws.row_dimensions[row_idx].height = row_height

    wb.save(out_xlsx)
    print("已保存 ->", out_xlsx)

if __name__ == "__main__":
    # 示例：自动缩放最长边到 200（可去掉 max_side 参数）
    image_to_excel("input.jpg", out_xlsx="pixel_art.xlsx", background="#000000", max_side=200)
